import math
import openpyxl

countFights = True
fightCount = 0
# Inputs
P1 = []
Pots1 = []
WL1 = []
WL2 = []
Pots2 = []
P2 = []

# Outputs
Players = []
PlayerPots = []
PlayerPotsUsed = []
PlayerWins = []
PlayerDraws = []
PlayerLosses = []
PFightAmount = []
PlayerElo = []

# Location of File
path = "pythonELO.xlsx"

# Create workbook object
wb = openpyxl.load_workbook(path)
fightsSheet = wb['Fights']
dataSheet = wb["Data"]
eloSheet = wb['Elo']

while countFights:
    if fightsSheet.cell(column=1, row=fightCount + 1).value is not None:
        fightCount += 1
    else:
        countFights = False

# Put all Data in Arrays
for i in range(fightCount):
    P1.append(fightsSheet.cell(column=1, row=i + 2).value)
    Pots1.append(fightsSheet.cell(column=2, row=i + 2).value)
    WL1.append(fightsSheet.cell(column=3, row=i + 2).value)
    WL2.append(fightsSheet.cell(column=4, row=i + 2).value)
    Pots2.append(fightsSheet.cell(column=5, row=i + 2).value)
    P2.append(fightsSheet.cell(column=6, row=i + 2).value)

# Make list of different players
for i in range(len(P1) - 1):
    # Add if the player is already in the array
    if P1[i] in Players:
        for j in range(len(Players)):
            if Players[j] == P1[i]:
                PlayerPots[j] += Pots1[i]
                PlayerPotsUsed[j] += (21 - Pots1[i])
                PFightAmount[j] += 1
                if WL1[i] == "W":
                    PlayerWins[j] += 1
                elif WL1[i] == "L":
                    PlayerLosses[j] += 1
                elif WL1[i] == "D":
                    PlayerDraws[j] += 1
                break
    # Make lists for different players
    elif P1[i] not in Players:
        Players.append(P1[i])
        PlayerElo.append(100)
        PlayerPots.append(Pots1[i])
        PlayerPotsUsed.append(21 - Pots1[i])
        PFightAmount.append(1)
        if WL1[i] == "W":
            PlayerWins.append(1)
            PlayerLosses.append(0)
            PlayerDraws.append(0)
        elif WL1[i] == "L":
            PlayerLosses.append(1)
            PlayerWins.append(0)
            PlayerDraws.append(0)
        elif WL1[i] == "D":
            PlayerLosses.append(0)
            PlayerWins.append(0)
            PlayerDraws.append(1)

# Make list of different players
for i in range(len(P2) - 1):
    # Add if the player is already in the array
    if P2[i] in Players:
        for j in range(len(Players)):
            if Players[j] == P2[i]:
                PlayerPots[j] += Pots2[i]
                PlayerPotsUsed[j] += (21 - Pots2[i])
                PFightAmount[j] += 1
                if WL2[i] == "W":
                    PlayerWins[j] += 1
                elif WL2[i] == "L":
                    PlayerLosses[j] += 1
                elif WL2[i] == "D":
                    PlayerDraws[j] += 1
                break
    # Make lists for different players
    elif P2[i] not in Players:
        Players.append(P2[i])
        PlayerElo.append(100)
        PlayerPots.append(Pots2[i])
        PlayerPotsUsed.append(21 - Pots2[i])
        PFightAmount.append(1)
        if WL2[i] == "W":
            PlayerWins.append(1)
            PlayerLosses.append(0)
            PlayerDraws.append(0)
        elif WL2[i] == "L":
            PlayerLosses.append(1)
            PlayerWins.append(0)
            PlayerDraws.append(0)
        elif WL2[i] == "D":
            PlayerLosses.append(0)
            PlayerWins.append(0)
            PlayerDraws.append(1)

# Output data to data sheet
for i in range(len(Players)):
    dataSheet.cell(column=1, row=i + 2).value = Players[i]  # Player
    dataSheet.cell(column=2, row=i + 2).value = PlayerWins[i]  # Wins
    dataSheet.cell(column=3, row=i + 2).value = PlayerDraws[i]  # Draws
    dataSheet.cell(column=4, row=i + 2).value = PlayerLosses[i]  # Losses
    dataSheet.cell(column=5, row=i + 2).value = PlayerPots[i]  # Total pots remaining
    dataSheet.cell(column=6, row=i + 2).value = PlayerPots[i] / PFightAmount[i]  # avg pots remaining
    dataSheet.cell(column=7, row=i + 2).value = PlayerPotsUsed[i]  # Total pots used
    dataSheet.cell(column=8, row=i + 2).value = PlayerPotsUsed[i] / PFightAmount[i]  # avg pots used

for i in range(len(P1)-1):
    for j in range(len(Players)):
        if Players[j] == P1[i]:
            for k in range(len(Players)):
                if Players[k] == P2[i]:
                    currEloJ = PlayerElo[j]
                    currEloK = PlayerElo[k]
                    if WL1[i] == "W":
                        PlayerElo[j] += currEloK/currEloJ * 10
                        PlayerElo[k] -= currEloK/currEloJ * 10
                    elif WL1[i] == "L":
                        PlayerElo[j] -= currEloJ/currEloK * 10
                        PlayerElo[k] += currEloJ/currEloK * 10

# Outputting Elo
for i in range(len(Players)):
    eloSheet.cell(column=1, row=i + 2).value = Players[i]
    eloSheet.cell(column=2, row=i + 2).value = round(PlayerElo[i])

wb.save("pythonELO.xlsx")
